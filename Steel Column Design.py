import numpy as np
import openpyxl as ex

i = 5
def Numbering():
    num = 1
    while num != 0:
        def Excel():
            IPE = ex.load_workbook('Steel Full IPE.xlsx')
            IPE = IPE.active
            Number = IPE.cell(row=i, column=1).value
            A = IPE.cell(row=i, column=10).value
            r_x = IPE.cell(row=i, column=15).value
            r_y = IPE.cell(row=i, column=19).value
            Fy = IPE.cell(row=i, column=22).value
            E = IPE.cell(row=i, column=24).value
            return Number, A, r_x, r_y, Fy, E

        # # Get the K value
        # def K():
        #     return 

        # Get the Fcr value
        def Get_Fcr():
            Excel_Data = Excel()
            A = float(Excel_Data[1])
            r_x = float(Excel_Data[2])
            r_y = float(Excel_Data[3])
            Fy = float(Excel_Data[4])
            E = float(Excel_Data[5])
            k_x= 1; k_y= 1; L= 5    # Should be delete!
            Landa_x = (k_x*L)/r_x
            Landa_y = (k_y*L)/r_y
            Landa = max(Landa_x,Landa_y)
            Fe=float((np.pi**2)*E)/((Landa)**2)   
            if Landa <= 139:
                fcr=(0.658**(Fy/Fe))*Fy
            else:
                fcr=0.877*Fe
            return fcr

        def check():
            Fcr_Data = Get_Fcr()
            Fcr = Fcr_Data[0]
            Excel_Data = Excel()
            A = Excel_Data[1]
            Pn=Fcr*A
            Pu = 10000    # Should be delete!
            if (Pu/(0.9*Pn)) > 1:   
                Conditions =  "Next IPE"
                # Return to Numbering and choose Higher one
                num = 1
            elif (Pu/(0.9*Pn)) < 0.85:
                Conditions = "Nothig"     
                # Return to Numbering and choose Lower one
                num = 1
            else:
                Conditions = "OK"
                num = 0
                # break
            return num, Conditions
        chek_Data = check()
        num = chek_Data[0]

    return 





