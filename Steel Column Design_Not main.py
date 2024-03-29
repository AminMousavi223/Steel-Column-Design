import numpy as np
import openpyxl as ex
import numpy as np
import pandas as pd

# input value

def input_value():
    # calculating pu with dead load and live load or directly
    try:
        D = float(input('Enter dead load (kg)'))
        L = float(input('Enter live load (kg)'))
        Pu = None
    except:
        D = None
        L = None

    if D is None or L is None:
        try:
            Pu = float(input('Enter  critical load (kg)'))
        except:
            Pu = None

    # combined loading
    def Load(D, L, Pu):
        if Pu == 0 or Pu == None:
            Pu1 = 1.4 * D
            Pu2 = 1.4 * D + 1.6 * L
            Pu = max(Pu1, Pu2)

        return Pu

    Pu = Load(D, L, Pu)
    Bay = 0

    # frame setting

    frame = input('Enter what is kind of your system ? 1D, 2D or 3D ?').lower()
    if frame in ['2d', '3d']:
        levels = int(input('How many levels exist ?'))

        if frame == '2d':
            Bay = [int(input('how many opening exist ?'))]
        else:
            Bay = [int(input('How many bay exist in xy dimension '))
                , int(input('How many bay exist in xz dimension '))]

    # length calculator
    def length_cal(frame):
        lengthColumn = []
        lengthOpening_xy = []
        lengthOpening_xz = []
        # length of column
        if frame == '1d':
            lengthColumn.append(float(input('enter length of column column (m) ')))
            return [lengthColumn]
        else:

            for i in range(1, levels + 1):
                lengthColumn.append(float(input(f'length of column level {i} (m)')))
            # length of opening
            for i in range(1, Bay[0] + 1):
                lengthOpening_xy.append(float(input(f'length of Bay in xy plate {i}'
                                                    f' (m) (start from left)')))
            if frame == '3d':
                for i in range(1, Bay[1] + 1):
                    lengthOpening_xz.append(float(input(f'length of Bay in xz plate {i}'
                                                        f' (m) (start from left)')))
                return [lengthColumn, lengthOpening_xy, lengthOpening_xz]

            return [lengthColumn, lengthOpening_xy]

    length = length_cal(frame)
    lengthColumn = length[0]
    if frame == '3d':
        lengthOpening = [length[1], length[2]]
    elif frame == '2d':
        lengthOpening = [length[1]]





    E = float(input('Enter E of your steel '))
    Fy = float(input('Enter Fy of your steel '))

    # K parameter or prerequisites
    try:
        k = float(input('Enter K if this parameter is clear ,otherwise just press enter'))
        if frame != '1d':
            return [frame, Pu, lengthColumn, E, Fy, lengthOpening , k]
        else:
            return [frame, Pu, lengthColumn, E, Fy, k]

    except:
        if frame == '1d':

            supportDown = input("what's your support in down of column ?")
            supportUp = input("what's your support in top of column ?")
            return [frame, Pu, lengthColumn, E, Fy, supportUp , supportDown]


        else:
            try:
                Ga = float(input('Enter Ga'))
                Gb = float(input('Enter Gb'))


            except:
                    columnNum_a = int(input('How many column is connect to point a'))
                    beamNum_a = int(input('How many beam is connect to point a'))
                    columnNum_b = int(input('How many column is connect to point b'))
                    beamNum_b = int(input('How many beam is connect to point b'))
                    Sigma = 0
                    sigma = []
                    for i in [columnNum_a , beamNum_a , columnNum_b , beamNum_b]:
                        for j in range(1,i+1) :
                            E = float(input(f'Enter E value of your {j} column or beam'))
                            I = float(input(f'Enter inertia moment of your {j} column or beam'))
                            L_joint = float(input(f'Enter lentgh of your {j} column or beam'))
                            Sigma += (E*I)/L_joint
                            sigma.append(Sigma)
                    Ga = sigma[0] / sigma[1]
                    Gb = sigma[2] / sigma[3]
            return [frame, Pu, lengthColumn, E, Fy, lengthOpening, Ga, Gb]

    return [frame, Pu, lengthColumn, E, Fy, lengthOpening]

input_value()



i = 5
def Numbering():
    num = 1
    while num != 0:
        def Excel():
            IPE = ex.load_workbook('Steel Full IPE.xlsx')
            IPE = IPE.active
            Number = IPE.cell(row=i, column=1).value
            A = IPE.cell(row=i, column=10).value
            r_x = IPE.cell(row=i, column=15).value
            r_y = IPE.cell(row=i, column=19).value
            Fy = IPE.cell(row=i, column=22).value
            E = IPE.cell(row=i, column=24).value
            return Number, A, r_x, r_y, Fy, E

        # # Get the K value
        # def K():
        #     return 

        # Get the Fcr value
        def Get_Fcr():
            Excel_Data = Excel()
            A = float(Excel_Data[1])
            r_x = float(Excel_Data[2])
            r_y = float(Excel_Data[3])
            Fy = float(Excel_Data[4])
            E = float(Excel_Data[5])
            k_x= 1; k_y= 1; L= 5    # Should be delete!
            Landa_x = (k_x*L)/r_x
            Landa_y = (k_y*L)/r_y
            Landa = max(Landa_x,Landa_y)
            Fe=float((np.pi**2)*E)/((Landa)**2)   
            if Landa <= 139:
                fcr=(0.658**(Fy/Fe))*Fy
            else:
                fcr=0.877*Fe
            return fcr

        def check():
            Fcr_Data = Get_Fcr()
            Fcr = Fcr_Data[0]
            Excel_Data = Excel()
            A = Excel_Data[1]
            Pn=Fcr*A
            Pu = 10000    # Should be delete!
            if (Pu/(0.9*Pn)) > 1:   
                Conditions =  "Next IPE"
                # Return to Numbering and choose Higher one
                num = 1
            elif (Pu/(0.9*Pn)) < 0.85:
                Conditions = "Nothig"     
                # Return to Numbering and choose Lower one
                num = 1
            else:
                Conditions = "OK"
                num = 0
                # break
            return num, Conditions
        chek_Data = check()
        num = chek_Data[0]

    return 





